"""
FASAL DOCTOR — Dataset Reviewer & Quality Checker
===================================================
Reviews extracted JSON records, flags missing fields,
lets you fix records interactively, and exports to
final Excel dataset.

Usage:
    python src/reviewer.py --input data/processed/extracted.json
    python src/reviewer.py --input data/processed/extracted.json --fix
    python src/reviewer.py --input data/processed/extracted.json --export
"""

import os, sys, json, argparse
from pathlib import Path
from colorama import Fore, Style, init

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

init(autoreset=True)

REQUIRED_FIELDS = [
    "id", "crop_name_en", "crop_name_ur", "disease_name_en", "disease_name_ur",
    "pathogen_type", "symptoms_en", "affected_part", "season",
    "severity_level", "spray_chemical_en", "brand_name_pakistan",
    "dose_per_acre", "spray_timing", "biological_control",
    "economic_loss_pct", "data_source", "verified_year", "country_code",
]
CRITICAL_FIELDS = [
    "crop_name_en", "disease_name_en", "symptoms_en",
    "spray_chemical_en", "severity_level",
]

def load_records(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def save_records(records: list, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

# ── Quality scoring ────────────────────────────────────────────────────────
def score_record(rec: dict) -> dict:
    def is_empty(v):
        if v is None: return True
        if isinstance(v, str): return not v.strip()
        return False
    missing_required = [f for f in REQUIRED_FIELDS if is_empty(rec.get(f))]
    missing_critical = [f for f in CRITICAL_FIELDS if is_empty(rec.get(f))]
    has_urdu         = bool(rec.get("crop_name_ur") or rec.get("disease_name_ur") or rec.get("symptoms_ur"))
    has_spray        = bool(rec.get("spray_chemical_en"))
    has_brand        = bool(rec.get("brand_name_pakistan"))
    has_dose         = bool(rec.get("dose_per_acre"))

    filled = len(REQUIRED_FIELDS) - len(missing_required)
    score  = round(filled / len(REQUIRED_FIELDS) * 100)

    return {
        "score":            score,
        "missing_required": missing_required,
        "missing_critical": missing_critical,
        "has_urdu":         has_urdu,
        "has_spray":        has_spray,
        "has_brand":        has_brand,
        "has_dose":         has_dose,
        "status":           "CRITICAL" if missing_critical else ("REVIEW" if score < 70 else "OK"),
    }

# ── Review report ──────────────────────────────────────────────────────────
def print_review(records: list):
    print(f"\n{'='*60}")
    print(f"{Fore.CYAN}DATASET QUALITY REVIEW{Style.RESET_ALL}")
    print(f"Total records: {len(records)}")
    print(f"{'='*60}\n")

    counts = {"OK": 0, "REVIEW": 0, "CRITICAL": 0}
    crop_counts  = {}
    low_conf     = []
    missing_brand= []
    missing_urdu = []

    for rec in records:
        q = score_record(rec)
        counts[q["status"]] += 1

        crop = rec.get("crop_name_en", "Unknown")
        crop_counts[crop] = crop_counts.get(crop, 0) + 1

        if rec.get("confidence") == "Low":
            low_conf.append(rec.get("id", "?"))
        if not q["has_brand"]:
            missing_brand.append(rec.get("id", "?"))
        if not q["has_urdu"]:
            missing_urdu.append(rec.get("id", "?"))

    # Status summary
    print(f"  {Fore.GREEN}OK       : {counts['OK']}{Style.RESET_ALL}")
    print(f"  {Fore.YELLOW}REVIEW   : {counts['REVIEW']}{Style.RESET_ALL}")
    print(f"  {Fore.RED}CRITICAL : {counts['CRITICAL']}{Style.RESET_ALL}")

    # Crops covered
    print(f"\n{Fore.CYAN}Crops covered ({len(crop_counts)}):{Style.RESET_ALL}")
    for crop, n in sorted(crop_counts.items(), key=lambda x: -x[1]):
        bar = "█" * min(n, 20)
        print(f"  {crop:<20} {bar} {n}")

    # Flags
    if low_conf:
        print(f"\n{Fore.YELLOW}Low confidence records ({len(low_conf)}): {', '.join(low_conf[:10])}{'...' if len(low_conf)>10 else ''}{Style.RESET_ALL}")
    if missing_brand:
        print(f"{Fore.YELLOW}Missing Pakistan brand ({len(missing_brand)}): {', '.join(missing_brand[:10])}{'...' if len(missing_brand)>10 else ''}{Style.RESET_ALL}")
    if missing_urdu:
        print(f"{Fore.YELLOW}Missing Urdu names ({len(missing_urdu)}): {', '.join(missing_urdu[:10])}{'...' if len(missing_urdu)>10 else ''}{Style.RESET_ALL}")

    # Detailed listing
    print(f"\n{Fore.CYAN}Record details:{Style.RESET_ALL}")
    for rec in records:
        q = score_record(rec)
        color = Fore.GREEN if q["status"]=="OK" else (Fore.YELLOW if q["status"]=="REVIEW" else Fore.RED)
        print(f"  {color}[{rec.get('id','?'):>7}] {q['score']:>3}% | "
              f"{rec.get('crop_name_en','?'):<12} | {rec.get('disease_name_en','?'):<35} | "
              f"{'Brand:Y' if q['has_brand'] else 'Brand:?'} "
              f"{'Urdu:Y' if q['has_urdu'] else 'Urdu:?'}{Style.RESET_ALL}")

    print(f"\n{'='*60}\n")

# ── Interactive fixer ──────────────────────────────────────────────────────
def interactive_fix(records: list) -> list:
    print(f"\n{Fore.CYAN}INTERACTIVE RECORD FIXER{Style.RESET_ALL}")
    print("Press Enter to skip a field, type new value to update, 'q' to quit\n")

    for i, rec in enumerate(records):
        q = score_record(rec)
        if q["status"] == "OK":
            continue

        print(f"\n{'─'*50}")
        print(f"{Fore.YELLOW}[{rec.get('id','?')}] {rec.get('crop_name_en')} — {rec.get('disease_name_en')}{Style.RESET_ALL}")
        print(f"Score: {q['score']}% | Status: {q['status']}")
        if q["missing_required"]:
            print(f"Missing: {', '.join(q['missing_required'])}")

        for field in q["missing_required"] + (["brand_name_pakistan"] if not q["has_brand"] else []):
            current = rec.get(field, "")
            val = input(f"  {field} [{current}]: ").strip()
            if val == "q":
                return records
            if val:
                records[i][field] = val

    return records

# ── Excel export ───────────────────────────────────────────────────────────
EXPORT_COLS = [
    "id", "crop_name_en", "crop_name_ur", "disease_name_en", "disease_name_ur",
    "pathogen_type", "symptoms_en", "symptoms_ur", "affected_part", "season",
    "region_pakistan", "climate_trigger", "severity_level",
    "spray_chemical_en", "brand_name_pakistan", "dose_per_acre", "spray_timing",
    "safety_precautions_ur", "biological_control", "economic_loss_pct",
    "data_source", "verified_year", "country_code", "confidence", "source_file",
]
SEV_COLORS = {"Critical":"FFCDD2","High":"FFE0B2","Medium":"FFF9C4","Low":"DCEDC8"}
PATH_COLORS = {"Fungal":"E8F5E9","Bacterial":"E3F2FD","Viral":"FCE4EC","Insect Pest":"FFF3E0"}

def export_excel(records: list, out_path: str):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Disease Database"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(EXPORT_COLS))}1")
    tc = ws["A1"]
    tc.value = "FASAL DOCTOR — Extracted Disease Database — Review Before Use"
    tc.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill  = PatternFill("solid", fgColor="1B5E20")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(EXPORT_COLS, start=1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2E7D32")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, rec in enumerate(records, start=3):
        q   = score_record(rec)
        bg  = "F1F8E9" if ri % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(EXPORT_COLS, start=1):
            val = rec.get(col, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = border
            if col == "severity_level":
                c.fill = PatternFill("solid", fgColor=SEV_COLORS.get(val, bg))
                c.font = Font(name="Arial", bold=True, size=9)
            elif col == "pathogen_type":
                c.fill = PatternFill("solid", fgColor=PATH_COLORS.get(val, bg))
            elif col == "id":
                c.fill = PatternFill("solid", fgColor="2E7D32")
                c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            else:
                c.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[ri].height = 60

    # Quality sheet
    ws2 = wb.create_sheet("Quality Report")
    ws2.sheet_view.showGridLines = False
    qh  = ["ID","Crop","Disease","Score%","Status","Missing Fields","Has Brand","Has Urdu"]
    for ci, h in enumerate(qh, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A237E")
        c.alignment = Alignment(horizontal="center")
    for ri, rec in enumerate(records, start=2):
        q  = score_record(rec)
        bg = "C8E6C9" if q["status"]=="OK" else ("FFF9C4" if q["status"]=="REVIEW" else "FFCDD2")
        row = [
            rec.get("id",""), rec.get("crop_name_en",""), rec.get("disease_name_en",""),
            q["score"], q["status"], ", ".join(q["missing_required"]),
            "Yes" if q["has_brand"] else "No", "Yes" if q["has_urdu"] else "No",
        ]
        for ci, val in enumerate(row, start=1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left")
    for ci in range(1, len(qh)+1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"{Fore.GREEN}Exported {len(records)} records → {out_path}{Style.RESET_ALL}")


def main():
    parser = argparse.ArgumentParser(description="Fasal Doctor Dataset Reviewer")
    parser.add_argument("--input",  required=True, help="Input JSON file")
    parser.add_argument("--fix",    action="store_true", help="Interactive fix mode")
    parser.add_argument("--export", action="store_true", help="Export to Excel")
    parser.add_argument("--out",    default="data/output/fasal_doctor_dataset.xlsx")
    args = parser.parse_args()

    records = load_records(args.input)
    print_review(records)

    if args.fix:
        records = interactive_fix(records)
        save_records(records, args.input)
        print(f"{Fore.GREEN}Saved fixed records → {args.input}{Style.RESET_ALL}")

    if args.export:
        export_excel(records, args.out)


if __name__ == "__main__":
    main()
